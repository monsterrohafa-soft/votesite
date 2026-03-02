#!/usr/bin/env python3
"""
SNS 검색 결과를 원본 후보자 데이터와 합쳐서 엑셀로 통합 저장
"""
import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = "/home/rohafa/sns_search"
SOURCE_FILE = "/home/rohafa/election_2026_전체후보자.json"
OUTPUT_FILE = "/home/rohafa/election_2026_전체후보자_SNS포함.xlsx"
DESKTOP_OUTPUT = "/mnt/c/Users/gyupe/Desktop/election_2026_전체후보자_SNS포함.xlsx"

PARTY_COLORS = {
    "더불어민주당": "DBEEF4",
    "국민의힘": "FCE4EC",
    "조국혁신당": "E8F5E9",
    "진보당": "FFF3E0",
    "개혁신당": "F3E5F5",
    "녹색정의당": "E8F5E9",
    "사회민주당": "FFF8E1",
    "무소속": "F5F5F5",
}

SNS_FIELDS = ["인스타그램", "페이스북", "블로그", "유튜브", "트위터", "홈페이지", "기타SNS"]

def clean_name(name):
    """성명에서 한자 괄호 제거: 김형남(金炯男) → 김형남"""
    return re.sub(r'\(.*?\)', '', name).strip()

def short_sido(sido):
    """시도명 축약: 서울특별시 → 서울, 경기도 → 경기"""
    mapping = {
        "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
        "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
        "울산광역시": "울산", "세종특별자치시": "세종", "경기도": "경기",
        "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남",
        "전북특별자치도": "전북", "전라남도": "전남", "경상북도": "경북",
        "경상남도": "경남", "제주특별자치도": "제주",
    }
    return mapping.get(sido, sido)

def load_sns_results():
    """모든 result_batch_*.json 로드하여 다중 키 딕셔너리 생성"""
    sns_map = {}
    for i in range(1, 18):
        path = os.path.join(BASE_DIR, f"result_batch_{i}.json")
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                for item in data:
                    name = clean_name(item.get("성명", ""))
                    party = item.get("소속정당명", "")
                    sido = short_sido(item.get("시도", ""))
                    sns_data = {}
                    for field in SNS_FIELDS:
                        val = item.get(field)
                        if val and val != "null":
                            sns_data[field] = val
                    if sns_data:
                        # 다중 키로 저장 (매칭률 높이기)
                        key1 = f"{name}__{party}__{sido}"
                        sns_map[key1] = sns_data
                        # 정당 없는 경우 (교육감 등) name+시도 키
                        key2 = f"{name}__*__{sido}"
                        sns_map[key2] = sns_data
                        # name+정당 키 (시도가 구이름인 경우)
                        key3 = f"{name}__{party}__*"
                        sns_map[key3] = sns_data
                        # name만 (최후 수단)
                        key4 = f"{name}__*__*"
                        sns_map[key4] = sns_data
                print(f"  batch_{i}: {len(data)}명 로드 완료")
            except Exception as e:
                print(f"  batch_{i}: 로드 실패 - {e}")
        else:
            print(f"  batch_{i}: 파일 없음 (미완료)")
    return sns_map

def find_sns(sns_map, name, party, sido):
    """다중 키 전략으로 SNS 매칭"""
    # 1순위: 이름+정당+시도 정확 매칭
    key1 = f"{name}__{party}__{sido}"
    if key1 in sns_map:
        return sns_map[key1]
    # 2순위: 이름+시도 (정당 무시, 교육감용)
    key2 = f"{name}__*__{sido}"
    if key2 in sns_map:
        return sns_map[key2]
    # 3순위: 이름+정당 (시도 무시, 구이름 문제용)
    key3 = f"{name}__{party}__*"
    if key3 in sns_map:
        return sns_map[key3]
    # 4순위: 이름만 (동명이인 위험 있지만 SNS 발견자가 적어서 OK)
    key4 = f"{name}__*__*"
    if key4 in sns_map:
        return sns_map[key4]
    return {}

def extract_candidates(source_data):
    """원본 JSON에서 선거유형별 후보자 목록 추출 (정규화)"""
    results = {}

    type_keys = {
        "시도지사선거": "시도지사",
        "교육감선거": "교육감",
        "구시군의장선거": "구시군의장",
        "시도의회의원선거": "시도의회의원",
        "구시군의회의원선거": "구시군의회의원",
    }

    for src_key, label in type_keys.items():
        if src_key not in source_data:
            continue
        section = source_data[src_key]
        candidates = []

        if "후보자목록" in section:
            # 구시군의장: 플랫 리스트
            for c in section["후보자목록"]:
                cand = {
                    "성명": clean_name(c.get("성명", "")),
                    "소속정당명": c.get("소속정당명", ""),
                    "시도": short_sido(c.get("시도", "")),
                    "선거구명": c.get("선거구명", ""),
                    "성별": c.get("성별", ""),
                    "생년월일": c.get("생년월일", ""),
                    "직업": c.get("직업", ""),
                    "학력": c.get("학력", ""),
                }
                candidates.append(cand)
        elif "시도별" in section:
            # 시도지사/교육감/시도의회/구시군의회: 시도별 → 리스트
            for sido_name, sido_list in section["시도별"].items():
                if isinstance(sido_list, list):
                    for c in sido_list:
                        cand = {
                            "성명": clean_name(c.get("성명", "")),
                            "소속정당명": c.get("소속정당명", ""),
                            "시도": short_sido(c.get("시도", sido_name)),
                            "선거구명": c.get("선거구명", ""),
                            "성별": c.get("성별", ""),
                            "생년월일": c.get("생년월일", ""),
                            "직업": c.get("직업", ""),
                            "학력": c.get("학력", ""),
                        }
                        candidates.append(cand)

        results[label] = candidates
        print(f"  {label}: {len(candidates)}명")

    return results

def create_excel(candidates_by_type, sns_map):
    """엑셀 파일 생성"""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    sheet_order = ["시도지사", "교육감", "구시군의장", "시도의회의원", "구시군의회의원"]
    stats = {}
    all_found = []

    for idx, etype in enumerate(sheet_order):
        if etype not in candidates_by_type:
            continue
        candidates = candidates_by_type[etype]
        if idx == 0:
            ws = wb.active
            ws.title = etype
        else:
            ws = wb.create_sheet(title=etype)

        base_cols = ["성명", "소속정당명", "시도", "선거구명", "성별", "생년월일", "직업"]
        all_cols = base_cols + SNS_FIELDS

        # 헤더
        for col_idx, col_name in enumerate(all_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        sns_count = 0
        for row_idx, cand in enumerate(candidates, 2):
            name = cand.get("성명", "")
            party = cand.get("소속정당명", "")
            sido = cand.get("시도", "")

            sns_data = find_sns(sns_map, name, party, sido)
            has_sns = bool(sns_data)
            if has_sns:
                sns_count += 1
                all_found.append({**cand, "선거유형": etype, **sns_data})

            row_data = [
                name, party, sido,
                cand.get("선거구명", ""),
                cand.get("성별", ""),
                cand.get("생년월일", ""),
                cand.get("직업", ""),
            ]
            for field in SNS_FIELDS:
                row_data.append(sns_data.get(field, ""))

            party_color = PARTY_COLORS.get(party, "FFFFFF")
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
                cell.border = thin_border
                cell.fill = PatternFill(start_color=party_color, end_color=party_color, fill_type="solid")
                if col_idx > len(base_cols) and value and isinstance(value, str) and value.startswith("http"):
                    cell.font = Font(color="0563C1", underline="single")
                    cell.hyperlink = value
                if has_sns and col_idx <= 3:
                    cell.font = Font(bold=True)

        # 열 너비
        widths = [12, 14, 10, 15, 5, 18, 15] + [35]*7
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(candidates)+1}"
        ws.freeze_panes = "A2"

        stats[etype] = {"count": len(candidates), "sns": sns_count}

    # === 요약 시트 ===
    ws_s = wb.create_sheet(title="요약", index=0)
    ws_s.merge_cells("A1:G1")
    cell = ws_s["A1"]
    cell.value = "2026 지방선거 예비후보자 SNS 검색 결과"
    cell.font = Font(bold=True, size=16, color="2F5496")
    cell.alignment = Alignment(horizontal="center")

    ws_s.cell(row=2, column=1, value="검색일: 2026-03-01 | batch 11 미완료 (200명)").font = Font(italic=True, color="888888")

    # 통계
    row = 4
    for col_idx, h in enumerate(["선거유형", "후보자 수", "SNS 발견", "발견율"], 1):
        c = ws_s.cell(row=row, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

    row = 5
    gt, gs = 0, 0
    for etype in sheet_order:
        if etype not in stats:
            continue
        s = stats[etype]
        rate_str = f"{s['sns']/s['count']*100:.1f}%" if s['count'] > 0 else "0%"
        for col_idx, val in enumerate([etype, s['count'], s['sns'], rate_str], 1):
            c = ws_s.cell(row=row, column=col_idx, value=val)
            c.border = thin_border; c.alignment = Alignment(horizontal="center")
        gt += s['count']; gs += s['sns']
        row += 1

    rate = f"{gs/gt*100:.1f}%" if gt > 0 else "0%"
    for col_idx, val in enumerate(["합계", gt, gs, rate], 1):
        c = ws_s.cell(row=row, column=col_idx, value=val)
        c.font = Font(bold=True); c.border = thin_border; c.alignment = Alignment(horizontal="center")

    # SNS 발견 목록
    row += 2
    ws_s.cell(row=row, column=1, value="SNS 발견된 후보자 목록").font = Font(bold=True, size=13, color="2F5496")
    row += 1

    found_h = ["성명", "선거유형", "소속정당명", "시도", "선거구명"] + SNS_FIELDS
    for col_idx, h in enumerate(found_h, 1):
        c = ws_s.cell(row=row, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.border = thin_border
    row += 1

    for item in all_found:
        vals = [item.get("성명",""), item.get("선거유형",""), item.get("소속정당명",""),
                item.get("시도",""), item.get("선거구명","")]
        for f in SNS_FIELDS:
            vals.append(item.get(f, ""))
        for col_idx, val in enumerate(vals, 1):
            c = ws_s.cell(row=row, column=col_idx, value=val if val else "")
            c.border = thin_border
            if val and isinstance(val, str) and val.startswith("http"):
                c.font = Font(color="0563C1", underline="single")
                c.hyperlink = val
        row += 1

    # 요약 열 너비
    for col_idx, w in enumerate([12, 14, 14, 10, 15] + [35]*7, 1):
        ws_s.column_dimensions[get_column_letter(col_idx)].width = w

    # 저장
    wb.save(OUTPUT_FILE)
    print(f"\n엑셀 저장: {OUTPUT_FILE}")

    try:
        import shutil
        shutil.copy2(OUTPUT_FILE, DESKTOP_OUTPUT)
        print(f" 바탕화면 복사: {DESKTOP_OUTPUT}")
    except Exception as e:
        print(f" 바탕화면 복사 실패: {e}")

    print(f"\n 최종: {gt}명 중 {gs}명 SNS 발견 ({gs/gt*100:.1f}%)")
    print(f" SNS 발견 목록:")
    for item in all_found:
        snss = [f for f in SNS_FIELDS if item.get(f)]
        print(f"   {item['성명']} ({item.get('소속정당명','')}, {item.get('시도','')} {item.get('선거구명','')}) → {', '.join(snss)}")

if __name__ == "__main__":
    print("=" * 60)
    print("2026 지방선거 후보자 SNS 통합 엑셀 생성")
    print("=" * 60)

    print("\n SNS 검색 결과 로드 중...")
    sns_map = load_sns_results()
    print(f"  SNS 발견 후보자: {len(sns_map)}명")

    print("\n 원본 후보자 데이터 로드 중...")
    with open(SOURCE_FILE, "r", encoding="utf-8") as f:
        source_data = json.load(f)
    candidates_by_type = extract_candidates(source_data)

    print("\n 엑셀 파일 생성 중...")
    create_excel(candidates_by_type, sns_map)

    print("\n 작업 완료!")
