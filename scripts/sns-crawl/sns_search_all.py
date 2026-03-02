#!/usr/bin/env python3
"""
2026 지방선거 예비후보자 SNS 전수 검색 스크립트
- Claude 토큰 소모 없이 독립 실행
- duckduckgo_search 라이브러리 사용 (API 키 불필요)
- 진행 상황 자동 저장 및 이어하기 지원
- 완료 후 엑셀 자동 생성

사용법:
  pip install duckduckgo_search openpyxl
  python3 sns_search_all.py

옵션:
  python3 sns_search_all.py --batch 11        # 특정 배치만 처리
  python3 sns_search_all.py --reprocess        # 이미 완료된 배치도 재처리
  python3 sns_search_all.py --merge-only       # 검색 없이 엑셀 통합만
  python3 sns_search_all.py --delay 3          # 검색 간 대기시간(초, 기본 2)
"""

import json
import os
import re
import sys
import time
import argparse
from datetime import datetime

# 출력 버퍼링 비활성화
sys.stdout.reconfigure(line_buffering=True)

# ============================================================
# 설정
# ============================================================
BASE_DIR = "/home/rohafa/sns_search"
SOURCE_FILE = "/home/rohafa/election_2026_전체후보자.json"
OUTPUT_EXCEL = "/home/rohafa/election_2026_전체후보자_SNS포함.xlsx"
DESKTOP_EXCEL = "/mnt/c/Users/gyupe/Desktop/election_2026_전체후보자_SNS포함.xlsx"
PROGRESS_FILE = os.path.join(BASE_DIR, "search_progress.json")

# SNS URL 패턴
SNS_PATTERNS = {
    "인스타그램": [
        r'https?://(?:www\.)?instagram\.com/([a-zA-Z0-9_.]+)',
    ],
    "페이스북": [
        r'https?://(?:www\.)?facebook\.com/([a-zA-Z0-9_.]+)',
        r'https?://(?:www\.)?fb\.com/([a-zA-Z0-9_.]+)',
    ],
    "블로그": [
        r'https?://blog\.naver\.com/([a-zA-Z0-9_]+)',
        r'https?://([a-zA-Z0-9_]+)\.tistory\.com',
        r'https?://brunch\.co\.kr/@([a-zA-Z0-9_]+)',
    ],
    "유튜브": [
        r'https?://(?:www\.)?youtube\.com/(?:channel|c|@)([a-zA-Z0-9_\-]+)',
        r'https?://(?:www\.)?youtube\.com/@([a-zA-Z0-9_\-]+)',
    ],
    "트위터": [
        r'https?://(?:www\.)?(?:twitter|x)\.com/([a-zA-Z0-9_]+)',
    ],
    "홈페이지": [
        r'(https?://[a-zA-Z0-9\-]+\.(?:kr|com|net|org)[^\s]*)',
    ],
}

# 무시할 URL (공식 사이트, 뉴스 등)
IGNORE_DOMAINS = [
    'nec.go.kr', 'naver.com/search', 'google.com', 'daum.net/search',
    'namu.wiki', 'wikipedia.org', 'namuwiki', 'news.', 'search.',
    'assembly.go.kr', 'council.', '.go.kr', 'dcinside.com',
    'youtube.com/watch', 'youtube.com/results',
]


# ============================================================
# 검색 엔진
# ============================================================
class DuckDuckGoSearcher:
    """DuckDuckGo 검색 (API 키 불필요)"""

    def __init__(self):
        try:
            from duckduckgo_search import DDGS
            self.ddgs = DDGS()
            self.available = True
            print("  [OK] DuckDuckGo 검색 엔진 준비 완료")
        except ImportError:
            self.available = False
            print("  [!] duckduckgo_search 미설치. 설치: pip install duckduckgo_search")

    def search(self, query, max_results=5):
        """검색 수행, 결과 URL+제목+본문 리스트 반환"""
        if not self.available:
            return []
        try:
            results = list(self.ddgs.text(query, region='kr-kr', max_results=max_results))
            return [
                {
                    "url": r.get("href", ""),
                    "title": r.get("title", ""),
                    "body": r.get("body", ""),
                }
                for r in results
            ]
        except Exception as e:
            print(f"    검색 오류: {e}")
            return []


class NaverSearcher:
    """네이버 검색 API (선택, API 키 필요)"""

    def __init__(self, client_id=None, client_secret=None):
        self.client_id = client_id or os.environ.get("NAVER_CLIENT_ID", "")
        self.client_secret = client_secret or os.environ.get("NAVER_CLIENT_SECRET", "")
        self.available = bool(self.client_id and self.client_secret)
        if self.available:
            print("  [OK] 네이버 검색 API 준비 완료")
        else:
            print("  [스킵] 네이버 API 키 미설정 (선택사항)")

    def search(self, query, max_results=5):
        import urllib.request
        import urllib.parse
        url = f"https://openapi.naver.com/v1/search/webkr.json?query={urllib.parse.quote(query)}&display={max_results}"
        req = urllib.request.Request(url)
        req.add_header("X-Naver-Client-Id", self.client_id)
        req.add_header("X-Naver-Client-Secret", self.client_secret)
        try:
            resp = urllib.request.urlopen(req, timeout=10)
            data = json.loads(resp.read().decode("utf-8"))
            return [
                {
                    "url": item.get("link", ""),
                    "title": re.sub(r'<[^>]+>', '', item.get("title", "")),
                    "body": re.sub(r'<[^>]+>', '', item.get("description", "")),
                }
                for item in data.get("items", [])
            ]
        except Exception as e:
            print(f"    네이버 검색 오류: {e}")
            return []


# ============================================================
# SNS 추출 로직
# ============================================================
def extract_sns_from_results(results, candidate_name):
    """검색 결과에서 SNS URL 추출"""
    found = {}

    for result in results:
        url = result.get("url", "")
        title = result.get("title", "")
        body = result.get("body", "")
        text = f"{url} {title} {body}"

        # URL 자체가 SNS 플랫폼인지 직접 확인 (가장 신뢰도 높음)
        url_lower = url.lower()

        # 인스타그램 - URL이 직접 인스타그램인 경우
        if "인스타그램" not in found and "instagram.com/" in url_lower:
            match = re.search(r'instagram\.com/([a-zA-Z0-9_.]+)', url)
            if match:
                username = match.group(1)
                if username.lower() not in ["p", "reel", "explore", "accounts", "stories", "tv"]:
                    found["인스타그램"] = f"https://www.instagram.com/{username}/"

        # 페이스북
        if "페이스북" not in found and ("facebook.com/" in url_lower or "fb.com/" in url_lower):
            match = re.search(r'facebook\.com/([a-zA-Z0-9_.]+)', url)
            if match:
                username = match.group(1)
                if username.lower() not in ["pages", "groups", "events", "watch", "share", "sharer", "photo", "login"]:
                    found["페이스북"] = f"https://www.facebook.com/{username}"

        # 트위터/X
        if "트위터" not in found and ("twitter.com/" in url_lower or "x.com/" in url_lower):
            match = re.search(r'(?:twitter|x)\.com/([a-zA-Z0-9_]+)', url)
            if match:
                username = match.group(1)
                if username.lower() not in ["home", "search", "explore", "hashtag", "i", "intent", "share", "login"]:
                    found["트위터"] = f"https://x.com/{username}"

        # 유튜브 채널
        if "유튜브" not in found and "youtube.com/" in url_lower:
            match = re.search(r'youtube\.com/(?:channel/|c/|@)([a-zA-Z0-9_\-%]+)', url)
            if match:
                found["유튜브"] = url

        # 블로그
        if "블로그" not in found:
            if "blog.naver.com/" in url_lower:
                found["블로그"] = url
            elif ".tistory.com" in url_lower:
                match = re.search(r'https?://([a-zA-Z0-9_-]+)\.tistory\.com', url)
                if match:
                    found["블로그"] = f"https://{match.group(1)}.tistory.com"
            elif "brunch.co.kr/@" in url_lower:
                found["블로그"] = url

        # 본문/제목에서 SNS URL 추가 추출 (이름이 언급된 경우만)
        if candidate_name in title or candidate_name in body:
            if "인스타그램" not in found:
                match = re.search(r'instagram\.com/([a-zA-Z0-9_.]+)', text)
                if match and match.group(1).lower() not in ["p", "reel", "explore"]:
                    found["인스타그램"] = f"https://www.instagram.com/{match.group(1)}/"

            if "트위터" not in found:
                match = re.search(r'(?:twitter|x)\.com/([a-zA-Z0-9_]+)', text)
                if match and match.group(1).lower() not in ["home", "search", "explore", "i", "intent"]:
                    found["트위터"] = f"https://x.com/{match.group(1)}"

    return found


def search_candidate(searchers, name, party, sido, district):
    """한 후보자에 대해 검색 수행"""
    # 다양한 검색 쿼리 (첫번째에서 SNS 찾으면 나머지 스킵)
    queries = [
        f'"{name}" {party} {district} site:instagram.com OR site:facebook.com OR site:youtube.com OR site:x.com OR site:blog.naver.com',
        f'"{name}" {party} {sido} SNS 인스타그램',
        f'"{name}" {party} {district}',
    ]

    all_results = []
    for searcher in searchers:
        if not searcher.available:
            continue
        for query in queries:
            try:
                results = searcher.search(query, max_results=5)
                all_results.extend(results)
            except Exception as e:
                print(f"    검색 오류: {e}")
                continue
            # SNS URL이 결과에 포함되어 있으면 추가 검색 불필요
            found = extract_sns_from_results(all_results, name)
            if found:
                return found
        break  # 첫 번째 검색엔진만 사용

    return extract_sns_from_results(all_results, name)


# ============================================================
# 배치 처리
# ============================================================
def load_progress():
    """진행 상황 로드"""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_progress(progress):
    """진행 상황 저장"""
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


def process_batch(batch_num, searchers, delay=2, reprocess=False):
    """배치 파일 처리"""
    batch_file = os.path.join(BASE_DIR, f"batch_{batch_num}.json")
    result_file = os.path.join(BASE_DIR, f"result_batch_{batch_num}.json")

    if not os.path.exists(batch_file):
        print(f"  [스킵] batch_{batch_num}.json 파일 없음")
        return

    # 이미 완료된 배치 확인
    if os.path.exists(result_file) and not reprocess:
        print(f"  [스킵] batch_{batch_num} 이미 완료 (재처리: --reprocess)")
        return

    with open(batch_file, "r", encoding="utf-8") as f:
        candidates = json.load(f)

    print(f"\n{'='*60}")
    print(f"  배치 {batch_num}: {len(candidates)}명 처리 시작")
    print(f"{'='*60}")

    # 진행 상황 로드 (이어하기)
    progress = load_progress()
    batch_key = f"batch_{batch_num}"
    processed_names = set(progress.get(batch_key, {}).get("processed", []))

    results = []
    sns_found_count = 0
    total = len(candidates)

    for idx, cand in enumerate(candidates):
        name = cand.get("성명", "")
        party = cand.get("소속정당명", "")
        sido = cand.get("시도", "")
        district = cand.get("선거구명", "")

        # 진행률 표시
        pct = (idx + 1) / total * 100
        status = f"  [{idx+1}/{total}] ({pct:.0f}%) {name} ({party}, {sido} {district})"

        # 이미 처리한 후보자 스킵
        unique_key = f"{name}_{party}_{district}"
        if unique_key in processed_names and not reprocess:
            # 기존 결과에서 가져오기
            result = {
                "성명": name,
                "선거유형": cand.get("선거유형", ""),
                "소속정당명": party,
                "시도": sido,
                "선거구명": district,
                "인스타그램": None, "페이스북": None, "블로그": None,
                "유튜브": None, "트위터": None, "홈페이지": None, "기타SNS": None,
            }
            results.append(result)
            print(f"{status} → 스킵 (이전 처리)")
            continue

        # 검색 수행
        sns_data = search_candidate(searchers, name, party, sido, district)

        result = {
            "성명": name,
            "선거유형": cand.get("선거유형", ""),
            "소속정당명": party,
            "시도": sido,
            "선거구명": district,
            "인스타그램": sns_data.get("인스타그램"),
            "페이스북": sns_data.get("페이스북"),
            "블로그": sns_data.get("블로그"),
            "유튜브": sns_data.get("유튜브"),
            "트위터": sns_data.get("트위터"),
            "홈페이지": sns_data.get("홈페이지"),
            "기타SNS": sns_data.get("기타SNS"),
        }
        results.append(result)

        if sns_data:
            sns_found_count += 1
            sns_list = ", ".join(f"{k}: {v}" for k, v in sns_data.items())
            print(f"{status} → 발견! {sns_list}")
        else:
            print(f"{status} → 없음")

        # 진행 상황 저장 (10명마다)
        processed_names.add(unique_key)
        if (idx + 1) % 10 == 0:
            if batch_key not in progress:
                progress[batch_key] = {}
            progress[batch_key]["processed"] = list(processed_names)
            progress[batch_key]["last_idx"] = idx
            progress[batch_key]["timestamp"] = datetime.now().isoformat()
            save_progress(progress)

        # 속도 제한
        time.sleep(delay)

    # 결과 저장
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    # 진행 상황 최종 저장
    if batch_key not in progress:
        progress[batch_key] = {}
    progress[batch_key]["processed"] = list(processed_names)
    progress[batch_key]["completed"] = True
    progress[batch_key]["timestamp"] = datetime.now().isoformat()
    progress[batch_key]["sns_found"] = sns_found_count
    save_progress(progress)

    print(f"\n  배치 {batch_num} 완료: {len(results)}명 처리, {sns_found_count}명 SNS 발견")
    print(f"  결과 저장: {result_file}")


# ============================================================
# 엑셀 통합 (merge_to_excel.py와 동일)
# ============================================================
def merge_to_excel():
    """모든 결과를 엑셀로 통합"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    PARTY_COLORS = {
        "더불어민주당": "DBEEF4", "국민의힘": "FCE4EC",
        "조국혁신당": "E8F5E9", "진보당": "FFF3E0",
        "개혁신당": "F3E5F5", "녹색정의당": "E8F5E9",
        "사회민주당": "FFF8E1", "무소속": "F5F5F5",
    }
    SNS_FIELDS = ["인스타그램", "페이스북", "블로그", "유튜브", "트위터", "홈페이지", "기타SNS"]

    def clean_name(name):
        return re.sub(r'\(.*?\)', '', name).strip()

    def short_sido(s):
        m = {
            "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
            "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
            "울산광역시": "울산", "세종특별자치시": "세종", "경기도": "경기",
            "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남",
            "전북특별자치도": "전북", "전라남도": "전남", "경상북도": "경북",
            "경상남도": "경남", "제주특별자치도": "제주",
        }
        return m.get(s, s)

    # SNS 결과 로드
    print("\n  SNS 결과 로드 중...")
    sns_map = {}
    for i in range(1, 18):
        path = os.path.join(BASE_DIR, f"result_batch_{i}.json")
        if not os.path.exists(path):
            continue
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            name = clean_name(item.get("성명", ""))
            party = item.get("소속정당명", "")
            sido = short_sido(item.get("시도", ""))
            sns_data = {}
            for field in SNS_FIELDS:
                val = item.get(field)
                if val and val != "null":
                    sns_data[field] = val
            if sns_data:
                for key in [
                    f"{name}__{party}__{sido}",
                    f"{name}__*__{sido}",
                    f"{name}__{party}__*",
                    f"{name}__*__*",
                ]:
                    sns_map[key] = sns_data

    def find_sns(name, party, sido):
        for key in [
            f"{name}__{party}__{sido}",
            f"{name}__*__{sido}",
            f"{name}__{party}__*",
            f"{name}__*__*",
        ]:
            if key in sns_map:
                return sns_map[key]
        return {}

    # 원본 데이터 로드
    print("  원본 데이터 로드 중...")
    with open(SOURCE_FILE, "r", encoding="utf-8") as f:
        source_data = json.load(f)

    type_keys = {
        "시도지사선거": "시도지사", "교육감선거": "교육감",
        "구시군의장선거": "구시군의장", "시도의회의원선거": "시도의회의원",
        "구시군의회의원선거": "구시군의회의원",
    }

    candidates_by_type = {}
    for src_key, label in type_keys.items():
        if src_key not in source_data:
            continue
        section = source_data[src_key]
        candidates = []
        if "후보자목록" in section:
            for c in section["후보자목록"]:
                candidates.append({
                    "성명": clean_name(c.get("성명", "")),
                    "소속정당명": c.get("소속정당명", ""),
                    "시도": short_sido(c.get("시도", "")),
                    "선거구명": c.get("선거구명", ""),
                    "성별": c.get("성별", ""), "생년월일": c.get("생년월일", ""),
                    "직업": c.get("직업", ""), "학력": c.get("학력", ""),
                })
        elif "시도별" in section:
            for sido_name, sido_list in section["시도별"].items():
                if isinstance(sido_list, list):
                    for c in sido_list:
                        candidates.append({
                            "성명": clean_name(c.get("성명", "")),
                            "소속정당명": c.get("소속정당명", ""),
                            "시도": short_sido(c.get("시도", sido_name)),
                            "선거구명": c.get("선거구명", ""),
                            "성별": c.get("성별", ""), "생년월일": c.get("생년월일", ""),
                            "직업": c.get("직업", ""), "학력": c.get("학력", ""),
                        })
        candidates_by_type[label] = candidates
        print(f"    {label}: {len(candidates)}명")

    # 엑셀 생성
    print("  엑셀 생성 중...")
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    sheet_order = ["시도지사", "교육감", "구시군의장", "시도의회의원", "구시군의회의원"]
    stats = {}
    all_found = []

    for idx, etype in enumerate(sheet_order):
        if etype not in candidates_by_type:
            continue
        candidates = candidates_by_type[etype]
        ws = wb.active if idx == 0 else wb.create_sheet(title=etype)
        if idx == 0:
            ws.title = etype

        base_cols = ["성명", "소속정당명", "시도", "선거구명", "성별", "생년월일", "직업"]
        all_cols = base_cols + SNS_FIELDS

        for col_idx, col_name in enumerate(all_cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = header_align; cell.border = thin_border

        sns_count = 0
        for row_idx, cand in enumerate(candidates, 2):
            name = cand["성명"]
            party = cand["소속정당명"]
            sido = cand["시도"]
            sns_data = find_sns(name, party, sido)
            has_sns = bool(sns_data)
            if has_sns:
                sns_count += 1
                all_found.append({**cand, "선거유형": etype, **sns_data})

            row_data = [name, party, sido, cand.get("선거구명",""),
                        cand.get("성별",""), cand.get("생년월일",""), cand.get("직업","")]
            for field in SNS_FIELDS:
                row_data.append(sns_data.get(field, ""))

            party_color = PARTY_COLORS.get(party, "FFFFFF")
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value if value else "")
                cell.border = thin_border
                cell.fill = PatternFill(start_color=party_color, end_color=party_color, fill_type="solid")
                if col_idx > len(base_cols) and value and isinstance(value, str) and value.startswith("http"):
                    cell.font = Font(color="0563C1", underline="single")
                    cell.hyperlink = value
                if has_sns and col_idx <= 3:
                    cell.font = Font(bold=True)

        widths = [12, 14, 10, 15, 5, 18, 15] + [35]*7
        for col_idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}{len(candidates)+1}"
        ws.freeze_panes = "A2"
        stats[etype] = {"count": len(candidates), "sns": sns_count}

    # 요약 시트
    ws_s = wb.create_sheet(title="요약", index=0)
    ws_s.merge_cells("A1:G1")
    cell = ws_s["A1"]
    cell.value = "2026 지방선거 예비후보자 SNS 검색 결과"
    cell.font = Font(bold=True, size=16, color="2F5496")
    cell.alignment = Alignment(horizontal="center")
    ws_s.cell(row=2, column=1, value=f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="888888")

    row = 4
    for col_idx, h in enumerate(["선거유형", "후보자 수", "SNS 발견", "발견율"], 1):
        c = ws_s.cell(row=row, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = thin_border

    row = 5
    gt, gs = 0, 0
    for etype in sheet_order:
        if etype not in stats: continue
        s = stats[etype]
        rate = f"{s['sns']/s['count']*100:.1f}%" if s['count'] > 0 else "0%"
        for col_idx, val in enumerate([etype, s['count'], s['sns'], rate], 1):
            c = ws_s.cell(row=row, column=col_idx, value=val)
            c.border = thin_border; c.alignment = Alignment(horizontal="center")
        gt += s['count']; gs += s['sns']; row += 1

    rate = f"{gs/gt*100:.1f}%" if gt > 0 else "0%"
    for col_idx, val in enumerate(["합계", gt, gs, rate], 1):
        c = ws_s.cell(row=row, column=col_idx, value=val)
        c.font = Font(bold=True); c.border = thin_border; c.alignment = Alignment(horizontal="center")

    # SNS 발견 목록
    row += 2
    ws_s.cell(row=row, column=1, value="SNS 발견된 후보자 목록").font = Font(bold=True, size=13, color="2F5496")
    row += 1
    found_h = ["성명", "선거유형", "소속정당명", "시도", "선거구명"] + SNS_FIELDS
    for col_idx, h in enumerate(found_h, 1):
        c = ws_s.cell(row=row, column=col_idx, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.border = thin_border
    row += 1
    for item in all_found:
        vals = [item.get("성명",""), item.get("선거유형",""), item.get("소속정당명",""),
                item.get("시도",""), item.get("선거구명","")]
        for f in SNS_FIELDS:
            vals.append(item.get(f, ""))
        for col_idx, val in enumerate(vals, 1):
            c = ws_s.cell(row=row, column=col_idx, value=val if val else "")
            c.border = thin_border
            if val and isinstance(val, str) and val.startswith("http"):
                c.font = Font(color="0563C1", underline="single")
                c.hyperlink = val
        row += 1

    for col_idx, w in enumerate([12, 14, 14, 10, 15] + [35]*7, 1):
        ws_s.column_dimensions[get_column_letter(col_idx)].width = w

    wb.save(OUTPUT_EXCEL)
    print(f"\n  엑셀 저장: {OUTPUT_EXCEL}")

    try:
        import shutil
        shutil.copy2(OUTPUT_EXCEL, DESKTOP_EXCEL)
        print(f"  바탕화면 복사: {DESKTOP_EXCEL}")
    except Exception as e:
        print(f"  바탕화면 복사 실패: {e}")

    print(f"\n  최종: {gt}명 중 {gs}명 SNS 발견 ({rate})")
    for item in all_found:
        snss = [f for f in SNS_FIELDS if item.get(f)]
        print(f"    {item['성명']} ({item.get('소속정당명','')}, {item.get('시도','')} {item.get('선거구명','')}) → {', '.join(snss)}")


# ============================================================
# 메인
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="2026 지방선거 후보자 SNS 전수 검색")
    parser.add_argument("--batch", type=int, help="특정 배치만 처리 (1~17)")
    parser.add_argument("--reprocess", action="store_true", help="이미 완료된 배치도 재처리")
    parser.add_argument("--merge-only", action="store_true", help="검색 없이 엑셀 통합만")
    parser.add_argument("--delay", type=float, default=2.0, help="검색 간 대기시간(초, 기본 2)")
    parser.add_argument("--naver-id", type=str, help="네이버 API Client ID")
    parser.add_argument("--naver-secret", type=str, help="네이버 API Client Secret")
    args = parser.parse_args()

    print("=" * 60)
    print("  2026 지방선거 예비후보자 SNS 전수 검색")
    print(f"  시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    if args.merge_only:
        print("\n[엑셀 통합만 실행]")
        merge_to_excel()
        return

    # 검색 엔진 초기화
    print("\n[검색 엔진 초기화]")
    searchers = []
    ddg = DuckDuckGoSearcher()
    if ddg.available:
        searchers.append(ddg)
    naver = NaverSearcher(args.naver_id, args.naver_secret)
    if naver.available:
        searchers.append(naver)

    if not searchers:
        print("\n사용 가능한 검색 엔진이 없습니다!")
        print("설치: pip install duckduckgo_search")
        print("또는 네이버 API 키 설정: --naver-id XXX --naver-secret YYY")
        sys.exit(1)

    # 배치 처리
    if args.batch:
        batches = [args.batch]
    else:
        batches = list(range(1, 18))

    print(f"\n[검색 시작] 배치: {batches}, 대기시간: {args.delay}초")

    for batch_num in batches:
        process_batch(batch_num, searchers, delay=args.delay, reprocess=args.reprocess)

    # 엑셀 통합
    print("\n" + "=" * 60)
    print("[엑셀 통합]")
    merge_to_excel()

    print("\n" + "=" * 60)
    print(f"  완료: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)


if __name__ == "__main__":
    main()
