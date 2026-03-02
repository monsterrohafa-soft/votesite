#!/usr/bin/env python3
"""
2026 지방선거 예비후보자 SNS 전수 검색 (Chrome 자동화 버전)
- 실제 크롬 브라우저로 네이버/구글 검색
- Selenium + webdriver_manager 사용
- 더블클릭으로 실행

사용법:
  pip install selenium webdriver-manager openpyxl
  python sns_search_chrome.py
"""

import json
import os
import re
import sys
import time
import argparse
from datetime import datetime

sys.stdout.reconfigure(line_buffering=True) if hasattr(sys.stdout, 'reconfigure') else None

# ============================================================
# 경로 설정 (스크립트 위치 기준 상대경로)
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_FILE = os.path.join(SCRIPT_DIR, "..", "election_2026_전체후보자.json")
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "..", "election_2026_전체후보자_SNS포함.xlsx")
PROGRESS_FILE = os.path.join(SCRIPT_DIR, "chrome_search_progress.json")

SNS_FIELDS = ["인스타그램", "페이스북", "블로그", "유튜브", "트위터", "홈페이지", "전화번호", "기타SNS"]


# ============================================================
# Chrome 브라우저 설정
# ============================================================
def detect_chrome_version():
    """설치된 Chrome의 메이저 버전 자동 감지"""
    import subprocess
    paths = [
        # Windows
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
    ]
    for p in paths:
        if os.path.exists(p):
            try:
                result = subprocess.run([p, "--version"], capture_output=True, text=True, timeout=5)
                ver = re.search(r'(\d+)\.\d+\.\d+\.\d+', result.stdout)
                if ver:
                    return int(ver.group(1))
            except Exception:
                pass
    # registry fallback
    try:
        result = subprocess.run(
            ['reg', 'query', r'HKEY_CURRENT_USER\Software\Google\Chrome\BLBeacon', '/v', 'version'],
            capture_output=True, text=True, timeout=5
        )
        ver = re.search(r'(\d+)\.\d+\.\d+\.\d+', result.stdout)
        if ver:
            return int(ver.group(1))
    except Exception:
        pass
    return None


def create_driver(headless=False):
    """Chrome WebDriver 생성 (구글 봇탐지 우회)"""
    try:
        import undetected_chromedriver as uc
        options = uc.ChromeOptions()
        if headless:
            options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--window-size=1280,900")
        options.add_argument("--lang=ko-KR")

        # Chrome 버전 자동 감지하여 맞는 ChromeDriver 사용
        chrome_ver = detect_chrome_version()
        if chrome_ver:
            print(f"  [i] Chrome version {chrome_ver} detected")
            driver = uc.Chrome(options=options, version_main=chrome_ver)
        else:
            print("  [i] Chrome version auto-detect failed, trying default...")
            driver = uc.Chrome(options=options)

        driver.implicitly_wait(3)
        print("  [OK] undetected-chromedriver (Google safe)")
        return driver
    except ImportError:
        print("  [!] undetected-chromedriver not found, using selenium")
    except Exception as e:
        print(f"  [!] undetected-chromedriver failed: {e}")
        print("  [i] Falling back to selenium...")

    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    options = webdriver.ChromeOptions()
    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1280,900")
    options.add_argument("--lang=ko-KR")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(3)
    print("  [OK] selenium + webdriver-manager")
    return driver


# ============================================================
# 네이버 검색
# ============================================================
def search_and_extract(driver, query, candidate_name, search_engine="google"):
    """검색 후 페이지 소스에서 직접 SNS URL 추출 (가장 정확)"""
    import urllib.parse

    if search_engine == "naver":
        url = f"https://search.naver.com/search.naver?query={urllib.parse.quote(query)}"
    else:
        url = f"https://www.google.com/search?q={urllib.parse.quote(query)}&hl=ko&gl=kr"

    try:
        driver.get(url)
        time.sleep(2)

        from selenium.webdriver.common.by import By

        found = {}

        # ========================================
        # 핵심: 페이지 소스에서 후보자 이름 근처의 SNS URL만 추출
        # ========================================
        source = driver.page_source

        # 이름 주변 1000자 내의 텍스트 조각들 수집
        nearby_texts = []
        for m in re.finditer(re.escape(candidate_name), source):
            chunk = source[max(0, m.start()-500):m.end()+500]
            nearby_texts.append(chunk)
        nearby_all = " ".join(nearby_texts)

        # 추가: a 태그에서 직접 SNS URL 수집 (빠른 방식)
        all_links = driver.find_elements(By.CSS_SELECTOR, "a[href]")
        sns_hrefs = []
        sns_domains = ["instagram.com", "facebook.com", "x.com", "twitter.com",
                       "youtube.com", "blog.naver.com", "tistory.com"]
        for link in all_links:
            href = link.get_attribute("href") or ""
            if any(d in href.lower() for d in sns_domains):
                text = link.text or ""
                sns_hrefs.append({"href": href, "text": text})

        # 이름 근처에 있는 SNS URL만 필터
        for item in sns_hrefs:
            href = item["href"]
            href_lower = href.lower()
            # href가 이름 근처 소스에 포함되어 있는지 확인
            name_nearby = (href in nearby_all) or (candidate_name in item["text"])

            # SNS 도메인 직접 매칭 (이름 근처 or 검색 결과 내)
            if "인스타그램" not in found and "instagram.com/" in href_lower:
                m = re.search(r'instagram\.com/([a-zA-Z0-9_.]+)', href)
                if m and m.group(1).lower() not in {"p","reel","reels","explore","accounts","stories","tv","about"}:
                    if name_nearby:
                        found["인스타그램"] = f"https://www.instagram.com/{m.group(1)}/"

            if "페이스북" not in found and "facebook.com/" in href_lower:
                m = re.search(r'facebook\.com/([a-zA-Z0-9_.]+)', href)
                if m and m.group(1).lower() not in {"pages","groups","events","watch","share","sharer","photo","login","help","policies"}:
                    if name_nearby:
                        found["페이스북"] = f"https://www.facebook.com/{m.group(1)}"

            if "트위터" not in found and ("twitter.com/" in href_lower or "x.com/" in href_lower):
                m = re.search(r'(?:twitter|x)\.com/([a-zA-Z0-9_]+)', href)
                if m and m.group(1).lower() not in {"home","search","explore","hashtag","i","intent","share","login","settings"}:
                    if name_nearby:
                        found["트위터"] = f"https://x.com/{m.group(1)}"

            if "유튜브" not in found and "youtube.com/" in href_lower:
                if re.search(r'youtube\.com/(?:channel/|@|c/)', href_lower):
                    if name_nearby:
                        found["유튜브"] = href

            if "블로그" not in found:
                if "blog.naver.com/" in href_lower and name_nearby:
                    m = re.search(r'blog\.naver\.com/([a-zA-Z0-9_]+)', href)
                    if m and m.group(1) not in {"MyBlog","PostList","PostView","BlogHome"}:
                        found["블로그"] = f"https://blog.naver.com/{m.group(1)}"
                elif ".tistory.com" in href_lower and name_nearby:
                    m = re.search(r'https?://([a-zA-Z0-9_\-]+)\.tistory\.com', href)
                    if m:
                        found["블로그"] = f"https://{m.group(1)}.tistory.com"

        # 2단계: 보이는 텍스트에서 전화번호 추출 (이름 근처)
        if "전화번호" not in found:
            visible = driver.find_element(By.TAG_NAME, "body").text
            for match in re.finditer(re.escape(candidate_name), visible):
                nearby = visible[max(0, match.start()-200):match.end()+300]
                # 한국 전화번호만: 02-xxxx-xxxx, 010-xxxx-xxxx, 031-xxx-xxxx 등
                phones = re.findall(r'(0(?:2|1[0-9]|[3-6][1-9])[-.\s]?\d{3,4}[-.\s]?\d{4})', nearby)
                if phones:
                    found["전화번호"] = phones[0].strip()
                    break

        return found

    except Exception as e:
        print(f"    검색 오류: {e}")
        return {}


# ============================================================
# SNS URL 추출
# ============================================================

def search_one_candidate(driver, name, party, sido, district, search_engine="google"):
    """한 후보자 검색"""
    # 이름에서 한자 제거: 김형남(金炯男) → 김형남
    clean = re.sub(r'\(.*?\)', '', name).strip()

    # 시도와 선거구가 같으면 중복 제거 (시도지사/교육감)
    location = district if district != sido else sido

    # 검색 쿼리 (+로 연결)
    queries = [
        f'{clean}+{party}+{location}+인스타그램+페이스북+SNS',
    ]

    for query in queries:
        found = search_and_extract(driver, query, clean, search_engine)
        if found:
            return found
        time.sleep(0.5)

    return {}


# ============================================================
# 진행 상황 관리
# ============================================================
def load_progress():
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_progress(progress):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)


# ============================================================
# 배치 처리
# ============================================================
def process_batch(driver, batch_num, search_engine="naver", delay=2, reprocess=False):
    """배치 파일 처리"""
    batch_file = os.path.join(SCRIPT_DIR, f"batch_{batch_num}.json")
    result_file = os.path.join(SCRIPT_DIR, f"result_batch_{batch_num}.json")

    if not os.path.exists(batch_file):
        print(f"  [SKIP] batch_{batch_num}.json not found")
        return

    if os.path.exists(result_file) and not reprocess:
        print(f"  [SKIP] batch_{batch_num} already done (use --reprocess)")
        return

    with open(batch_file, "r", encoding="utf-8") as f:
        candidates = json.load(f)

    print(f"\n{'='*60}")
    print(f"  Batch {batch_num}: {len(candidates)} candidates")
    print(f"{'='*60}")

    progress = load_progress()
    batch_key = f"batch_{batch_num}"
    done_keys = set(progress.get(batch_key, {}).get("done", []))

    # 기존 결과 로드 (이어하기용)
    existing_results = {}
    if os.path.exists(result_file) and not reprocess:
        try:
            with open(result_file, "r", encoding="utf-8") as f:
                for item in json.load(f):
                    k = f"{item['성명']}_{item.get('선거구명','')}"
                    existing_results[k] = item
        except:
            pass

    results = []
    sns_found = 0
    total = len(candidates)

    for idx, cand in enumerate(candidates):
        name = cand.get("성명", "")
        party = cand.get("소속정당명", "")
        sido = cand.get("시도", "")
        district = cand.get("선거구명", "")
        unique = f"{name}_{district}"

        clean = re.sub(r'\(.*?\)', '', name).strip()
        location = district if district != sido else sido
        pct = (idx + 1) / total * 100
        prefix = f"  [{idx+1}/{total}] ({pct:.0f}%) {clean} ({party}, {location})"

        # 이미 처리된 후보
        if unique in done_keys and not reprocess:
            if unique in existing_results:
                results.append(existing_results[unique])
            else:
                results.append({
                    "성명": name, "선거유형": cand.get("선거유형", ""),
                    "소속정당명": party, "시도": sido, "선거구명": district,
                    "인스타그램": None, "페이스북": None, "블로그": None,
                    "유튜브": None, "트위터": None, "홈페이지": None, "전화번호": None, "기타SNS": None,
                })
            print(f"{prefix} -> SKIP (done)")
            continue

        # 검색
        try:
            sns = search_one_candidate(driver, name, party, sido, district, search_engine)
        except Exception as e:
            print(f"{prefix} -> ERROR: {e}")
            sns = {}

        result = {
            "성명": name,
            "선거유형": cand.get("선거유형", ""),
            "소속정당명": party,
            "시도": sido,
            "선거구명": district,
            "인스타그램": sns.get("인스타그램"),
            "페이스북": sns.get("페이스북"),
            "블로그": sns.get("블로그"),
            "유튜브": sns.get("유튜브"),
            "트위터": sns.get("트위터"),
            "홈페이지": sns.get("홈페이지"),
            "전화번호": sns.get("전화번호"),
            "기타SNS": sns.get("기타SNS"),
        }
        results.append(result)

        if sns:
            sns_found += 1
            items = ", ".join(f"{k}={v}" for k, v in sns.items())
            print(f"{prefix} -> FOUND! {items}")
        else:
            print(f"{prefix} -> none")

        done_keys.add(unique)

        # 10명마다 저장
        if (idx + 1) % 10 == 0:
            with open(result_file, "w", encoding="utf-8") as f:
                json.dump(results, f, ensure_ascii=False, indent=2)
            progress[batch_key] = {
                "done": list(done_keys),
                "last": idx,
                "time": datetime.now().isoformat(),
            }
            save_progress(progress)

        time.sleep(delay)

    # 최종 저장
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    progress[batch_key] = {
        "done": list(done_keys),
        "completed": True,
        "time": datetime.now().isoformat(),
        "sns_found": sns_found,
    }
    save_progress(progress)

    print(f"\n  Batch {batch_num} done: {len(results)} processed, {sns_found} SNS found")


# ============================================================
# 엑셀 통합
# ============================================================
def merge_to_excel():
    """결과를 엑셀로 통합"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    PARTY_COLORS = {
        "더불어민주당": "DBEEF4", "국민의힘": "FCE4EC",
        "조국혁신당": "E8F5E9", "진보당": "FFF3E0",
        "개혁신당": "F3E5F5", "녹색정의당": "E8F5E9",
        "사회민주당": "FFF8E1", "무소속": "F5F5F5",
    }

    def clean_name(n):
        return re.sub(r'\(.*?\)', '', n).strip()

    def short_sido(s):
        m = {
            "서울특별시": "서울", "부산광역시": "부산", "대구광역시": "대구",
            "인천광역시": "인천", "광주광역시": "광주", "대전광역시": "대전",
            "울산광역시": "울산", "세종특별자치시": "세종", "경기도": "경기",
            "강원특별자치도": "강원", "충청북도": "충북", "충청남도": "충남",
            "전북특별자치도": "전북", "전라남도": "전남", "경상북도": "경북",
            "경상남도": "경남", "제주특별자치도": "제주",
        }
        return m.get(s, s)

    # SNS 결과 로드
    print("\n  Loading SNS results...")
    sns_map = {}
    for i in range(1, 18):
        path = os.path.join(SCRIPT_DIR, f"result_batch_{i}.json")
        if not os.path.exists(path):
            continue
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        for item in data:
            name = clean_name(item.get("성명", ""))
            party = item.get("소속정당명", "")
            sido = short_sido(item.get("시도", ""))
            sns_data = {}
            for field in SNS_FIELDS:
                val = item.get(field)
                if val and val != "null":
                    sns_data[field] = val
            if sns_data:
                for key in [
                    f"{name}__{party}__{sido}",
                    f"{name}__*__{sido}",
                    f"{name}__{party}__*",
                    f"{name}__*__*",
                ]:
                    sns_map[key] = sns_data

    def find_sns(name, party, sido):
        for key in [
            f"{name}__{party}__{sido}",
            f"{name}__*__{sido}",
            f"{name}__{party}__*",
            f"{name}__*__*",
        ]:
            if key in sns_map:
                return sns_map[key]
        return {}

    # 원본 데이터
    print("  Loading source data...")
    with open(SOURCE_FILE, "r", encoding="utf-8") as f:
        src = json.load(f)

    type_keys = {
        "시도지사선거": "시도지사", "교육감선거": "교육감",
        "구시군의장선거": "구시군의장", "시도의회의원선거": "시도의회의원",
        "구시군의회의원선거": "구시군의회의원",
    }

    cands_by_type = {}
    for src_key, label in type_keys.items():
        if src_key not in src:
            continue
        section = src[src_key]
        cands = []
        if "후보자목록" in section:
            for c in section["후보자목록"]:
                cands.append({
                    "성명": clean_name(c.get("성명", "")),
                    "소속정당명": c.get("소속정당명", ""),
                    "시도": short_sido(c.get("시도", "")),
                    "선거구명": c.get("선거구명", ""),
                    "성별": c.get("성별", ""),
                    "생년월일": c.get("생년월일", ""),
                    "직업": c.get("직업", ""),
                })
        elif "시도별" in section:
            for sn, sl in section["시도별"].items():
                if isinstance(sl, list):
                    for c in sl:
                        cands.append({
                            "성명": clean_name(c.get("성명", "")),
                            "소속정당명": c.get("소속정당명", ""),
                            "시도": short_sido(c.get("시도", sn)),
                            "선거구명": c.get("선거구명", ""),
                            "성별": c.get("성별", ""),
                            "생년월일": c.get("생년월일", ""),
                            "직업": c.get("직업", ""),
                        })
        cands_by_type[label] = cands
        print(f"    {label}: {len(cands)}")

    # 엑셀 생성
    print("  Creating Excel...")
    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    order = ["시도지사", "교육감", "구시군의장", "시도의회의원", "구시군의회의원"]
    stats = {}
    all_found = []

    for idx, et in enumerate(order):
        if et not in cands_by_type:
            continue
        cands = cands_by_type[et]
        ws = wb.active if idx == 0 else wb.create_sheet(title=et)
        if idx == 0:
            ws.title = et

        bcols = ["성명", "소속정당명", "시도", "선거구명", "성별", "생년월일", "직업"]
        acols = bcols + SNS_FIELDS

        for ci, cn in enumerate(acols, 1):
            c = ws.cell(row=1, column=ci, value=cn)
            c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border

        sc = 0
        for ri, cand in enumerate(cands, 2):
            nm, pt, sd = cand["성명"], cand["소속정당명"], cand["시도"]
            sns = find_sns(nm, pt, sd)
            has = bool(sns)
            if has:
                sc += 1
                all_found.append({**cand, "선거유형": et, **sns})

            rd = [nm, pt, sd, cand.get("선거구명",""), cand.get("성별",""),
                  cand.get("생년월일",""), cand.get("직업","")]
            for f in SNS_FIELDS:
                rd.append(sns.get(f, ""))

            pc = PARTY_COLORS.get(pt, "FFFFFF")
            for ci, val in enumerate(rd, 1):
                c = ws.cell(row=ri, column=ci, value=val if val else "")
                c.border = border
                c.fill = PatternFill(start_color=pc, end_color=pc, fill_type="solid")
                if ci > len(bcols) and val and isinstance(val, str) and val.startswith("http"):
                    c.font = Font(color="0563C1", underline="single")
                    c.hyperlink = val
                if has and ci <= 3:
                    c.font = Font(bold=True)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(acols))}{len(cands)+1}"
        ws.freeze_panes = "A2"
        for ci, w in enumerate([12,14,10,15,5,18,15]+[35]*7, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        stats[et] = {"count": len(cands), "sns": sc}

    # 요약 시트
    ws_s = wb.create_sheet(title="요약", index=0)
    ws_s.merge_cells("A1:G1")
    c = ws_s["A1"]
    c.value = "2026 지방선거 예비후보자 SNS 검색 결과"
    c.font = Font(bold=True, size=16, color="2F5496")
    c.alignment = Alignment(horizontal="center")
    ws_s.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="888888")

    row = 4
    for ci, h in enumerate(["선거유형", "후보자 수", "SNS 발견", "발견율"], 1):
        c = ws_s.cell(row=row, column=ci, value=h)
        c.font = hfont; c.fill = hfill; c.alignment = halign; c.border = border
    row = 5; gt = gs = 0
    for et in order:
        if et not in stats: continue
        s = stats[et]
        r = f"{s['sns']/s['count']*100:.1f}%" if s['count'] > 0 else "0%"
        for ci, v in enumerate([et, s['count'], s['sns'], r], 1):
            c = ws_s.cell(row=row, column=ci, value=v)
            c.border = border; c.alignment = Alignment(horizontal="center")
        gt += s['count']; gs += s['sns']; row += 1
    r = f"{gs/gt*100:.1f}%" if gt > 0 else "0%"
    for ci, v in enumerate(["합계", gt, gs, r], 1):
        c = ws_s.cell(row=row, column=ci, value=v)
        c.font = Font(bold=True); c.border = border; c.alignment = Alignment(horizontal="center")

    row += 2
    ws_s.cell(row=row, column=1, value="SNS 발견 후보자").font = Font(bold=True, size=13, color="2F5496")
    row += 1
    fh = ["성명", "선거유형", "소속정당명", "시도", "선거구명"] + SNS_FIELDS
    for ci, h in enumerate(fh, 1):
        c = ws_s.cell(row=row, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        c.border = border
    row += 1
    for item in all_found:
        vs = [item.get("성명",""), item.get("선거유형",""), item.get("소속정당명",""),
              item.get("시도",""), item.get("선거구명","")]
        for f in SNS_FIELDS:
            vs.append(item.get(f, ""))
        for ci, v in enumerate(vs, 1):
            c = ws_s.cell(row=row, column=ci, value=v if v else "")
            c.border = border
            if v and isinstance(v, str) and v.startswith("http"):
                c.font = Font(color="0563C1", underline="single"); c.hyperlink = v
        row += 1
    for ci, w in enumerate([12,14,14,10,15]+[35]*7, 1):
        ws_s.column_dimensions[get_column_letter(ci)].width = w

    wb.save(OUTPUT_EXCEL)
    print(f"\n  Excel saved: {OUTPUT_EXCEL}")
    print(f"  Total: {gt} candidates, {gs} SNS found ({r})")
    for item in all_found:
        snss = [f for f in SNS_FIELDS if item.get(f)]
        print(f"    {item['성명']} ({item.get('소속정당명','')}) -> {', '.join(snss)}")


# ============================================================
# 메인
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="SNS Search with Chrome")
    parser.add_argument("--batch", type=int, help="Process specific batch (1-17)")
    parser.add_argument("--reprocess", action="store_true", help="Re-process completed batches")
    parser.add_argument("--merge-only", action="store_true", help="Excel merge only")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay between searches (sec)")
    parser.add_argument("--headless", action="store_true", help="Run Chrome headless (no window)")
    parser.add_argument("--engine", default="google", choices=["naver", "google"], help="Search engine")
    args = parser.parse_args()

    print("=" * 60)
    print("  2026 Election - SNS Search (Chrome)")
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Engine: {args.engine}")
    print("=" * 60)

    if args.merge_only:
        merge_to_excel()
        return

    # Chrome 시작
    print("\n  Starting Chrome...")
    driver = create_driver(headless=args.headless)
    print("  Chrome ready!")

    try:
        batches = [args.batch] if args.batch else list(range(1, 18))
        print(f"  Batches: {batches}")
        print(f"  Delay: {args.delay}s")

        for bn in batches:
            process_batch(driver, bn, args.engine, args.delay, args.reprocess)

    finally:
        driver.quit()
        print("\n  Chrome closed.")

    # 엑셀 통합
    print("\n" + "=" * 60)
    print("  Excel merge...")
    merge_to_excel()

    print("\n" + "=" * 60)
    print(f"  Finished: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)


if __name__ == "__main__":
    main()
